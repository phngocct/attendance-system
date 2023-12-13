import os
from datetime import datetime
import pickle
import tkinter
import tkinter.messagebox
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO

import customtkinter
import face_recognition
from PIL import Image, ImageDraw
import numpy as np
import pandas as pd
import cv2
import openpyxl
import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
from firebase_admin import storage


def connect_database():
    cred = credentials.Certificate("serviceAccountKey.json")
    firebase_admin.initialize_app(
        cred,
        {
            "databaseURL": "https://faceattendancesystem-a35f4-default-rtdb.asia-southeast1.firebasedatabase.app/",
            "storageBucket": "faceattendancesystem-a35f4.appspot.com",
        },
    )
    ref = db.reference("Students")
    # Download file encode gương mặt từ Firebase Storage
    # filename = 'EncodeFile.pkl'
    # bucket = storage.bucket()
    # blob = bucket.blob(filename)
    # blob.download_to_filename(filename)
    return ref


def resize_25(image_path):
    # Đọc ảnh
    image = cv2.imread(image_path)
    # Tính kích thước mới
    width = int(image.shape[1] * 0.25)
    height = int(image.shape[0] * 0.25)
    # Resize ảnh
    resized_image = cv2.resize(image, (width, height))
    # Lưu ảnh đã resize
    resize_folder = "resize"
    if not os.path.exists(resize_folder):
        os.makedirs(resize_folder)
    folder_path, file_name = os.path.split(image_path)
    filename = file_name.split(".")[0]
    cv2.imwrite(os.path.join(resize_folder, "{}.png".format(filename)), resized_image)
    image_resize_path = os.path.join(resize_folder, "{}.png".format(filename))

    return image_resize_path


def add_data_to_database(student_id, name, major, image_path_100, ref):
    result_folder = "results_hog"
    face_detection_model = "hog"
    # result_folder = "results_cnn"
    # face_detection_model = "cnn"

    if not os.path.exists(result_folder):
        os.makedirs(result_folder)
    print(
        "---------------------------------- start -------------------------------------------"
    )
    start = datetime.now()
    image_path = resize_25(image_path_100)

    class Student:
        def __init__(self, student_id, name, major):
            self.student_id = student_id
            self.name = name
            self.major = major

        def to_dict(self):
            return {
                "name": self.name,
                "major": self.major,
                "total_attendance": 0,
                "last_attendance_time": "0000-00-00 00:00:00",
            }

    # Thêm ảnh vừa upload Firebase Storage
    def upload_image_to_storage(image_path):
        new_file_name = f"Images_{face_detection_model}/{student_id}.png"
        bucket = storage.bucket()
        blob = bucket.blob(new_file_name)
        blob.upload_from_filename(image_path)
        # Trả về đường dẫn ảnh trên Firebase Storage
        return blob.public_url

    def face_detections(image_path):
        # Đọc ảnh chứa khuôn mặt
        image_array = face_recognition.load_image_file(image_path)
        # Tìm kiếm vị trí khuôn mặt trong ảnh
        face_locations = face_recognition.face_locations(
            image_array, model=face_detection_model
        )

        return face_locations

    # Thêm ảnh cắt khuôn mặt lên Firebase Storage
    def crop_face(image_path, face_locations, output_size):
        print(f"{name}'s face location: ", face_locations)
        # Kiểm tra nếu không tìm thấy khuôn mặt
        if len(face_locations) == 0:
            print("Không tìm thấy khuôn mặt trong ảnh.")
        # Lấy tọa độ của khuôn mặt đầu tiên
        top, right, bottom, left = face_locations[0]
        # Tải hình ảnh
        image = Image.open(image_path)
        face_image = image.crop((left, top, right, bottom))
        resized_face_image = face_image.resize(output_size)
        # Lưu hình ảnh khuôn mặt đã resize
        output_folder = f"faces_{face_detection_model}"
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        resized_face_image.save(
            os.path.join(output_folder, "{}.png".format(student_id))
        )
        # Upload ảnh gương mặt lên Firebase Storage
        new_file_name = f"Faces_{face_detection_model}/{student_id}.png"
        face_image_path = f"faces_{face_detection_model}/{student_id}.png"
        bucket = storage.bucket()
        blob = bucket.blob(new_file_name)
        blob.upload_from_filename(face_image_path)
        # Trả về đường dẫn ảnh trên Firebase Storage
        return blob.public_url

    # Encode khuôn mặt lên Firebase Storage
    def encodings(image_path, face_locations):
        # Đọc ảnh chứa khuôn mặt
        image_array = face_recognition.load_image_file(image_path)
        # Mã hóa các khuôn mặt
        face_encodings = face_recognition.face_encodings(image_array, face_locations)
        # print('face_encodings: ', face_encodings)
        # print('type face_encodings: ', type(face_encodings))
        encodeKnownWithIds = [student_id, face_encodings]
        filename = f"EncodeFile_{face_detection_model}.pkl"
        if not os.path.exists(filename):
            # Tạo và định dạng file pickle
            encodeKnownWithIds = [encodeKnownWithIds]
            with open(filename, "wb") as file:
                pickle.dump(encodeKnownWithIds, file)
            file.close()

        else:
            # Đọc nội dung hiện có từ EncodeFile.pkl vào một biến
            with open(filename, "rb") as file:
                existing_data = pickle.load(file)
            # Thêm nội dung mới vào biến đã đọc
            existing_data.append(encodeKnownWithIds)
            # Ghi biến chứa nội dung mới vào file .pkl
            with open(filename, "wb") as file:
                pickle.dump(existing_data, file)
            file.close()
        # Upload file encode gương mặt lên Firebase Storage
        bucket = storage.bucket()
        blob = bucket.blob(filename)
        blob.upload_from_filename(filename)
        return blob.public_url

    def save_face_locations_xlsx(face_locations):
        if len(face_locations) == 0:
            top, right, bottom, left = 0, 0, 0, 0
        else:
            top, right, bottom, left = face_locations[0]

        xlsx_folder = "xlsx"
        if not os.path.exists(xlsx_folder):
            os.makedirs(xlsx_folder)
        # Kiểm tra nếu file Excel chưa tồn tại
        filename = f"face_locations_{face_detection_model}.xlsx"
        if not os.path.isfile(os.path.join(xlsx_folder, filename)):
            filename = f"face_locations_{face_detection_model}.xlsx"
            # Tạo file Excel mới
            workbook = openpyxl.Workbook()
            # Chọn sheet mặc định
            sheet = workbook.active
            # Tạo danh sách các heading
            headings = ["Student ID", "Top", "Right", "Bottom", "Left"]
            # Ghi các heading vào hàng đầu tiên của sheet
            for col_num, heading in enumerate(headings, start=1):
                sheet.cell(row=1, column=col_num).value = heading
            # Lưu file Excel
            workbook.save(os.path.join(xlsx_folder, filename))

        workbook = openpyxl.load_workbook(os.path.join(xlsx_folder, filename))
        sheet = workbook.active
        # Xác định hàng cuối cùng có dữ liệu trong sheet
        last_row = sheet.max_row + 1
        print("last_row: ", last_row)
        # Thêm dữ liệu vào từng ô trong sheet
        sheet.cell(row=last_row, column=1, value=student_id)
        sheet.cell(row=last_row, column=2, value=top)
        sheet.cell(row=last_row, column=3, value=right)
        sheet.cell(row=last_row, column=4, value=bottom)
        sheet.cell(row=last_row, column=5, value=left)
        # Lưu file Excel sau khi thêm dữ liệu
        workbook.save(os.path.join(xlsx_folder, filename))
        print("Add face locations data to excel file successfully")
        print("----------------------------------------------------------------------")

    def on_task_done(future):
        # Được gọi khi một công việc hoàn thành
        print(f"Công việc hoàn thành, kết quả: {future.result()}")

    # Tạo một đối tượng sinh viên
    student = Student(student_id=student_id, name=name, major=major)
    # Chuyển đổi đối tượng sinh viên thành từ điển
    student_dict = student.to_dict()
    # Thêm dữ liệu vào Firebase
    ref.child(student.student_id).set(student_dict)

    # Tạo một ThreadPoolExecutor với 2 luồng
    executor1 = ThreadPoolExecutor(max_workers=2)
    # Thực thi hàm trong luồng và truyền đối số bằng lambda function
    future1 = executor1.submit(lambda: upload_image_to_storage(image_path))
    future2 = executor1.submit(lambda: face_detections(image_path))
    # Thêm callback khi công việc hoàn thành
    future1.add_done_callback(on_task_done)
    future2.add_done_callback(on_task_done)
    # Chờ tất cả các công việc hoàn thành
    executor1.shutdown(wait=True)
    print(
        "Công việc của luồng upload_image_to_storage và luồng face_detections hoàn thành"
    )
    # Lấy kết quả trả về khi thực thi hoàn thành
    face_locations = future2.result()

    # Tạo một ThreadPoolExecutor với 2 luồng
    executor2 = ThreadPoolExecutor(max_workers=2)
    # Thực thi hàm trong luồng và truyền đối số bằng lambda function
    future3 = executor2.submit(
        lambda: crop_face(image_path, face_locations, (216, 216))
    )
    future4 = executor2.submit(lambda: encodings(image_path, face_locations))
    # Thêm callback khi công việc hoàn thành
    future3.add_done_callback(on_task_done)
    future4.add_done_callback(on_task_done)
    # Chờ tất cả các công việc hoàn thành
    executor2.shutdown(wait=True)
    print("Công việc của luồng crop_face và luồng encodings hoàn thành")
    # Lưu tọa độ của gương mặt sinh viên
    # save_face_locations_xlsx(face_locations)

    # Hiển thị thông báo thành công.
    # tkinter.messagebox.showinfo("Success", "Add new student successfully.")

    duration = datetime.now() - start
    time_completed = str(duration)
    data_eval = (
        "Student ID: "
        + str(student_id)
        + "\n"
        + "Full name: "
        + str(name)
        + "\n"
        + "Major: "
        + str(major)
        + "\n"
        + "Face locations: "
        + str(face_locations)
        + "\n"
        + "Time completed: "
        + str(duration)
    )
    print(data_eval)

    eval_path = f"time_use_{face_detection_model}_{student_id}.txt"
    path_s = os.path.join(result_folder, eval_path)
    with open(path_s, mode="w", encoding="utf-8") as f:
        f.writelines(data_eval)
    f.close()

    # Tạo DataFrame từ các danh sách
    data = {
        "Student ID": [student_id],
        "Full name": [name],
        "Major": [major],
        "Face locations": [face_locations],
        "Time completed": [time_completed],
    }
    df = pd.DataFrame(data)
    # Tạo index
    index = pd.Index(range(len(student_id)))
    # Lưu DataFrame vào file Excel
    xlsx_path = f"time_use_{face_detection_model}_{student_id}.xlsx"
    xlsx_path_s = os.path.join(result_folder, xlsx_path)
    df.to_excel(xlsx_path_s, index=False)

    return time_completed


def load_encode_file(filename):
    # Đọc nội dung từ file 'EncodeFile.pkl'
    with open(filename, "rb") as file:
        encodeListKnownWithIds = pickle.load(file)
    file.close()

    stdList = []
    # facelocationsList = []
    encodefaceList = []
    # Hiển thị danh sách face_encodings và student_id
    for i in range(len(encodeListKnownWithIds)):
        encodeList = encodeListKnownWithIds[i]
        std_id = encodeList[0]
        # face_locations = encodeList[1]
        encode_face = encodeList[1]
        stdList.append(std_id)
        # facelocationsList.append(face_locations)
        encodefaceList.append(encode_face)

    # Chuyển danh sách thành np.array
    # encodefaceList = np.array(encodefaceList)
    # Giảm số chiều thành 2
    # encodefaceList = np.squeeze(encodefaceList)

    # print("EncodefaceList: ", encodefaceList)

    print("Students List: ", stdList)
    print("Total student: ", len(stdList))
    # print("lenght facelocationsList: ", len(facelocationsList))
    print("lenght face_encodings: ", len(encodefaceList))
    print(
        "------------------------------------------------------------------------------------------------"
    )
    return stdList, encodefaceList


def recognize_result(
    face_encoding_to_check, known_face_encoding, studentIds, tolerance=0.4
):
    threshold = tolerance
    known_face_encoding = np.array(known_face_encoding)
    known_face_encoding = np.squeeze(known_face_encoding)

    face_encoding_to_check = np.array(face_encoding_to_check)
    matches = face_recognition.compare_faces(
        known_face_encoding, face_encoding_to_check, tolerance=threshold
    )
    faceDis = face_recognition.face_distance(
        known_face_encoding, face_encoding_to_check
    )
    match = any(matches)  # True/False
    count = matches.count(True)
    matchIndex = np.argmin(faceDis)
    distance = faceDis[matchIndex]
    distance = round(distance, 5)
    stdId = studentIds[matchIndex]

    print("----------------------------------------------------------------------")
    print("Matches: \n", matches)
    print("Match: ", match)
    print("Count True: ", count)
    print("Face distance: \n", faceDis)
    print("Min match Index: ", matchIndex)
    print("Face distance[ minmatchIndex] ", distance)
    print("----------------------------------------------------------------------")
    if match & (distance < threshold):
        print("Student ID: ", stdId)
        # Tính IOU
        iou_score = (1 / (1 + distance)) *100
        iou_score = round(iou_score, 5)
        # known_box = known_face_locations[matchIndex][0]
        # print(known_box)
        # box_to_check = face_locations_to_check[0]
        # iou_score = iou(known_box, box_to_check)
        print("----------------------------------------------------------------------")
        print("iou_score: ", iou_score)

        return stdId, distance, iou_score
    else:
        print("Không tìm thấy khuôn mặt khớp")
        return None, None, None


def get_data_to_database(stdId):
    studentInfo = db.reference(f"Students/{stdId}").get()
    print(studentInfo)
    name = studentInfo["name"]
    major = studentInfo["major"]
    total_attendance = studentInfo["total_attendance"]
    last_attendance_time = studentInfo["last_attendance_time"]

    return name, major, total_attendance, last_attendance_time


def update_data_to_database(stdId):
    ref = db.reference(f"Students/{stdId}")
    studentInfo = db.reference(f"Students/{stdId}").get()
    studentInfo["total_attendance"] += 1
    ref.child("total_attendance").set(studentInfo["total_attendance"])
    ref.child("last_attendance_time").set(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


def get_face_locations(image_arr_path):
    face_locations = face_recognition.face_locations(image_arr_path, model="hog")
    print("I found {} face(s) in this webcam.".format(len(face_locations)))
    print("Face location in webcam: ", face_locations)
    print(
        "------------------------------------------------------------------------------------------------"
    )
    if not face_locations:
        return None
    return face_locations


def get_image_from_db(stdId):
    bucket = storage.bucket()
    blob = bucket.blob(f"Images_hog/{stdId}.png")
    image_data = blob.download_as_bytes()
    # Chuyển đổi ảnh từ dạng bytes sang đối tượng PIL Image
    image = Image.open(BytesIO(image_data))

    return image


def draw_bounding_box(image, face_locations, stdId, distance):
    for face_location in face_locations:
        top, right, bottom, left = face_location
        top, right, bottom, left = top * 4, right * 4, bottom * 4, left * 4
        # Vẽ bounding box
        cv2.rectangle(image, (left, top), (right, bottom), (0, 255, 0), 2)

        # Hiển thị stdId và distance trên bounding box
        if stdId is None:
            label = " Unknown"
            cv2.rectangle(
                image, (left, top - 50), (right, top), (0, 255, 0), cv2.FILLED
            )
        else:
            label = f" {stdId}  {distance:.3f} %"
            # Tính độ dài của văn bản
            text_size, _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.9, 2)
            text_width = text_size[0]
            text_width = int(text_width / 2)
            cv2.rectangle(
                image,
                (left, top - 50),
                (right + text_width, top),
                (0, 255, 0),
                cv2.FILLED,
            )

        cv2.putText(
            image, label, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 0), 2
        )

    return image


def resize_image(image, min_size, max_size):
    width, height = image.size
    # Tìm đường chéo của ảnh
    diagonal = (width**2 + height**2) ** 0.5
    # Tính toán tỷ lệ giữa kích thước tối thiểu và đường chéo
    scale = min_size / diagonal
    # Kiểm tra nếu tỷ lệ vượt quá kích thước tối đa
    if diagonal * scale > max_size:
        scale = max_size / diagonal
    # Tính toán kích thước mới dựa trên tỷ lệ
    new_width = int(width * scale)
    new_height = int(height * scale)
    # Thay đổi kích thước của ảnh
    resized_image = image.resize((new_width, new_height), Image.ANTIALIAS)

    return resized_image


def add_data_to_excel(file_path, last_attendance_time, student_id, name, major):
    log_folder = "log"
    if not os.path.exists(log_folder):
        os.makedirs(log_folder)
    # Kiểm tra nếu file Excel chưa tồn tại
    if not os.path.isfile(os.path.join(log_folder, file_path)):
        # Tạo tên file Excel dựa trên ngày tháng năm hiện tại
        now = datetime.now()
        file_name = f"log_{now.strftime('%d-%m-%Y')}.xlsx"
        # Tạo file Excel mới
        workbook = openpyxl.Workbook()
        # Chọn sheet mặc định
        sheet = workbook.active
        # Tạo danh sách các heading
        headings = ["Time", "Student ID", "Full Name", "Major"]
        # Ghi các heading vào hàng đầu tiên của sheet
        for col_num, heading in enumerate(headings, start=1):
            sheet.cell(row=1, column=col_num).value = heading
        # Lưu file Excel
        workbook.save(os.path.join(log_folder, file_path))

    workbook = openpyxl.load_workbook(os.path.join(log_folder, file_path))
    sheet = workbook.active
    # Xác định hàng cuối cùng có dữ liệu trong sheet
    last_row = sheet.max_row + 1
    # print("last_row: ", last_row)
    # Thêm dữ liệu vào từng ô trong sheet
    sheet.cell(row=last_row, column=1, value=last_attendance_time)
    sheet.cell(row=last_row, column=2, value=student_id)
    sheet.cell(row=last_row, column=3, value=name)
    sheet.cell(row=last_row, column=4, value=major)
    # Lưu file Excel sau khi thêm dữ liệu
    workbook.save(os.path.join(log_folder, file_path))
    print("Add attendance data to excel file successfully")
    print("----------------------------------------------------------------------")
    # Upload file Excel lên Firebase Storage
    file_name = f"{log_folder}/{file_path}"
    bucket = storage.bucket()
    blob = bucket.blob(file_name)
    blob.upload_from_filename(file_name)
    # Trả về đường dẫn ảnh trên Firebase Storage
    return blob.public_url


def load_excel(file_path):
    log_folder = "log"
    if not os.path.exists(log_folder):
        os.makedirs(log_folder)
    # Kiểm tra nếu file Excel chưa tồn tại
    if not os.path.isfile(os.path.join(log_folder, file_path)):
        # Tạo tên file Excel dựa trên ngày tháng năm hiện tại
        now = datetime.now()
        file_name = f"log_{now.strftime('%d-%m-%Y')}.xlsx"
        # Tạo file Excel mới
        workbook = openpyxl.Workbook()
        # Chọn sheet mặc định
        sheet = workbook.active
        # Tạo danh sách các heading
        headings = ["Time", "Student ID", "Full Name", "Major"]
        # Ghi các heading vào hàng đầu tiên của sheet
        for col_num, heading in enumerate(headings, start=1):
            sheet.cell(row=1, column=col_num).value = heading
        # Lưu file Excel
        workbook.save(os.path.join(log_folder, file_path))
    workbook = openpyxl.load_workbook(os.path.join(log_folder, file_path))
    sheet = workbook.active

    list_values = list(sheet.values)
    cols = list_values[0]
    value_tuples = list_values[1:]
    value_tuples.reverse()

    return cols, value_tuples


def load_data_from_firebase():
    student_data = db.reference("Students").get()
    std_ids, names, majors, total_attendances, last_attendance_times = (
        [],
        [],
        [],
        [],
        [],
    )
    if student_data is not None:
        for std_id, std_info in student_data.items():
            name = std_info["name"]
            major = std_info["major"]
            total_attendance = std_info["total_attendance"]
            last_attendance_time = std_info["last_attendance_time"]
            std_ids.append(std_id),
            names.append(name),
            majors.append(major),
            total_attendances.append(total_attendance),
            last_attendance_times.append(last_attendance_time)
            # Thực hiện xử lý dữ liệu tùy ý ở đây
            # print(f"Student ID: {std_id}")
            # print(f"Name: {name}")
            # print(f"Major: {major}")
            # print(f"Total Attendance: {total_attendance}")
            # print(f"Last Attendance Time: {last_attendance_time}")
            # print("--------------------")

        return std_ids, names, majors, total_attendances, last_attendance_times
    else:
        print("No student data available")
        return None, None, None, None, None


def delete_data_from_firebase(std_id):
    ref = db.reference(f"Students/{std_id}")
    # Xóa dữ liệu
    ref.delete()

    face_detection_model = "hog"
    file_images_path = f"Images_{face_detection_model}/{std_id}.png"
    file_faces_path = f"Faces_{face_detection_model}/{std_id}.png"
    bucket = storage.bucket()
    file_images_ref = bucket.blob(file_images_path)
    file_faces_ref = bucket.blob(file_faces_path)
    # Xóa tệp tin
    file_images_ref.delete()
    file_faces_ref.delete()
    print("Successfully deleted student data on FireBase.")


def delete_std_from_encode(_id, filename):
    # Đọc nội dung từ file 'EncodeFile_hog.pkl'
    with open(filename, "rb") as file:
        encodeListKnownWithIds = pickle.load(file)
    # Tạo danh sách encodefaceList và stdList
    stdList = []
    # facelocationsList = []
    encodefaceList = []

    for encodeList in encodeListKnownWithIds:
        std_id = encodeList[0]
        # face_locations = encodeList[1]
        encode_face = encodeList[1]
        stdList.append(std_id)
        # facelocationsList.append(face_locations)
        encodefaceList.append(encode_face)

    try:
        # Tìm vị trí của std_id trong stdList
        index = stdList.index(_id)
        # Xóa phần tử tương ứng từ cả hai danh sách
        stdList.pop(index)
        # facelocationsList.pop(index)
        encodefaceList.pop(index)
        print(f"Xóa thành công sinh viên có std_id: {_id}")
    except ValueError:
        print(f"Sinh viên với std_id {_id} không tồn tại trong danh sách.")

    # Ghi lại dữ liệu vào file
    with open(filename, "wb") as file:
        pickle.dump(list(zip(stdList, encodefaceList)), file)

    print("Students List: ", stdList)
    print("Total student: ", len(stdList))
    # print("lenght facelocationsList: ", len(facelocationsList))
    print("lenght face_encodings: ", len(encodefaceList))
    print(
        "------------------------------------------------------------------------------------------------"
    )
    print("Successfully deleted student data from encodefile.")


def iou(box1, box2):
    """
    box1 -- hình hộp thứ nhất, đối tượng danh sách chứa tọa độ (top, right, bottom, left)
    box2 -- hình hộp thứ hai, đối tượng danh sách chứa tọa độ (top, right, bottom, left)
    """

    top1, right1, bottom1, left1 = box1
    top2, right2, bottom2, left2 = box2

    xi1 = max(left1, left2)
    yi1 = max(top1, top2)
    xi2 = min(right1, right2)
    yi2 = min(bottom1, bottom2)
    inter_area = max(0, yi2 - yi1) * max(0, xi2 - xi1)

    box1_area = (bottom1 - top1) * (right1 - left1)
    box2_area = (bottom2 - top2) * (right2 - left2)
    union_area = box1_area + box2_area - inter_area

    # Tính chỉ số IoU
    iou_score = inter_area / union_area

    iou_score = iou_score * 100
    iou_score = round(iou_score, 3)

    return iou_score
